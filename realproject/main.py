import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.graphics import Color, Rectangle
import calendar
import datetime
import openpyxl
from openpyxl import Workbook
import os
from functools import partial


class LoginPage(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = [20, 50, 20, 20]  # left, top, right, bottom
        self.spacing = 20

        self.add_widget(Label(text="Enter PIN:", font_size='24sp'))
        self.pin_input = TextInput(password=True, multiline=False, size_hint_y=None, height='40dp')
        self.add_widget(self.pin_input)

        login_button = Button(text="Login", font_size='24sp', size_hint_y=None, height='50dp', on_release=self.check_pin)
        self.add_widget(login_button)

    def check_pin(self, instance):
        if self.pin_input.text == "2256":
            self.parent.manager.current = 'main'
        else:
            popup = Popup(title='Error', content=Label(text='Incorrect PIN'), size_hint=(0.6, 0.4))
            popup.open()


class MainPage(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = [20, 20, 20, 20]
        self.spacing = 20

        self.file_name = 'work_hours.xlsx'
        self.hours_worked = {}
        self.current_day = None
        self.current_year = datetime.datetime.now().year
        self.current_month = datetime.datetime.now().month

        self.create_file()
        self.load_minimum_wage()

        self.add_widget(Label(text="Multition Pay", font_size='32sp', bold=True, size_hint_y=None, height='50dp'))

        self.total_label = Label(text="Total Salary: KWN 0.00", font_size='24sp', size_hint_y=None, height='40dp')
        self.add_widget(self.total_label)

        button_layout = BoxLayout(size_hint_y=None, height='50dp', spacing=20)
        set_wage_button = Button(text="Set Minimum Wage", on_release=self.set_minimum_wage)
        next_month_button = Button(text="Next Month", on_release=self.next_month)
        button_layout.add_widget(set_wage_button)
        button_layout.add_widget(next_month_button)
        self.add_widget(button_layout)

        self.calendar_layout = GridLayout(cols=7, padding=[0, 20, 0, 0], spacing=5)
        self.add_widget(self.calendar_layout)

        self.load_calendar()

    def create_file(self):
        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Work Hours'
            ws.append(['Year', 'Month', 'Day', 'Hours'])
            ws = wb.create_sheet(title='Settings')
            ws.append(['Minimum Wage', 0])
            wb.save(self.file_name)
        else:
            wb = openpyxl.load_workbook(self.file_name)
            if 'Settings' not in wb.sheetnames:
                ws = wb.create_sheet(title='Settings')
                ws.append(['Minimum Wage', 0])
            if 'Work Hours' not in wb.sheetnames:
                ws = wb.create_sheet(title='Work Hours')
                ws.append(['Year', 'Month', 'Day', 'Hours'])
            wb.save(self.file_name)

    def load_minimum_wage(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Settings']
        self.minimum_wage = ws['B1'].value
        print(f"Loaded minimum wage: {self.minimum_wage}")
        if self.minimum_wage == 0:
            self.set_minimum_wage()
        wb.close()

    def set_minimum_wage(self, instance=None):
        def save_wage(instance):
            try:
                wage = float(wage_input.text)
                self.minimum_wage = wage
                wb = openpyxl.load_workbook(self.file_name)
                ws = wb['Settings']
                ws['B1'] = self.minimum_wage
                wb.save(self.file_name)
                self.calculate_salary()
                popup.dismiss()
                print(f"Saved minimum wage: {self.minimum_wage}")
            except ValueError:
                popup.content = Label(text='Please enter a valid number')

        layout = BoxLayout(orientation='vertical', padding=20, spacing=20)
        layout.add_widget(Label(text="Enter your minimum wage per hour:", font_size='18sp'))
        wage_input = TextInput(multiline=False, size_hint_y=None, height='40dp')
        layout.add_widget(wage_input)
        save_button = Button(text="Save", size_hint_y=None, height='50dp', on_release=save_wage)
        layout.add_widget(save_button)

        popup = Popup(title='Set Minimum Wage', content=layout, size_hint=(0.8, 0.6))
        popup.open()

    def load_calendar(self):
        self.calendar_layout.clear_widgets()
        cal = calendar.TextCalendar(calendar.SUNDAY)
        month_days = cal.monthdayscalendar(self.current_year, self.current_month)

        days_labels = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        for day_label in days_labels:
            self.calendar_layout.add_widget(Label(text=day_label, font_size='18sp'))

        for week in month_days:
            for day in week:
                if day == 0:
                    self.calendar_layout.add_widget(Label(text=""))
                else:
                    day_box = BoxLayout(orientation='vertical', padding=5)
                    with day_box.canvas.before:
                        today = datetime.datetime.now()
                        if self.current_year == today.year and self.current_month == today.month and day == today.day:
                            Color(1, 1, 0, 1)  # Yellow color
                            self.rect = Rectangle(size=day_box.size, pos=day_box.pos)
                            day_box.bind(size=self._update_rect, pos=self._update_rect)

                    date_label = Label(text=str(day), font_size='18sp')
                    day_box.add_widget(date_label)

                    entry = TextInput(multiline=False, size_hint_y=None, height='40dp')
                    entry.bind(focus=partial(self.on_entry_focus, day))
                    day_box.add_widget(entry)

                    self.load_hours(day, entry)
                    self.calendar_layout.add_widget(day_box)

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def on_entry_focus(self, day, instance, value):
        if not value:  # Focus lost
            text = instance.text
            if text == "":
                if day in self.hours_worked:
                    del self.hours_worked[day]
                self.delete_hours(day)
            else:
                try:
                    hours = float(text)
                    self.hours_worked[day] = hours
                    self.save_hours(day, hours)
                except ValueError:
                    pass
            self.calculate_salary()

    def calculate_salary(self):
        total_hours = sum(self.hours_worked.values())
        print(f"Total hours: {total_hours}")
        total_salary = total_hours * self.minimum_wage
        print(f"Total salary: {total_salary}")
        self.total_label.text = f"Total Salary: KWN {total_salary:.2f}"

    def save_hours(self, day, hours):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Work Hours']
        found = False
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == self.current_year and row[1].value == self.current_month and row[2].value == day:
                row[3].value = hours
                found = True
                break
        if not found:
            ws.append([self.current_year, self.current_month, day, hours])
        wb.save(self.file_name)
        print(f"Saved hours for day {day}: {hours}")

    def load_hours(self, day, entry):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Work Hours']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == self.current_year and row[1] == self.current_month and row[2] == day:
                entry.text = str(row[3])
                self.hours_worked[day] = float(row[3])
                break
        self.calculate_salary()
        wb.close()
        print(f"Loaded hours for day {day}: {entry.text}")

    def delete_hours(self, day):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Work Hours']
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == self.current_year and row[1].value == self.current_month and row[2].value == day:
                ws.delete_rows(row[0].row, 1)
                break
        wb.save(self.file_name)
        print(f"Deleted hours for day {day}")

    def next_month(self, instance):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.hours_worked.clear()
        self.load_calendar()
        self.calculate_salary()


class MultitionPayApp(App):
    def build(self):
        sm = ScreenManager()
        login_screen = Screen(name='login')
        login_screen.add_widget(LoginPage())
        sm.add_widget(login_screen)

        main_screen = Screen(name='main')
        main_screen.add_widget(MainPage())
        sm.add_widget(main_screen)

        return sm


if __name__ == "__main__":
    MultitionPayApp().run()
